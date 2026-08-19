import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

source = Path(sys.argv[1] if len(sys.argv) > 1 else "outputs/mhd-catalog-products.csv")
destination = Path(sys.argv[2] if len(sys.argv) > 2 else "outputs/mhd-catalog-products.xlsx")

with source.open(newline="", encoding="utf-8") as file:
    rows = list(csv.reader(file, delimiter=";"))

workbook = Workbook()
sheet = workbook.active
sheet.title = "Productos MHD"
for row in rows:
    sheet.append(row)

sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
table = Table(displayName="ProductosMHD", ref=sheet.dimensions)
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
sheet.add_table(table)
for cell in sheet[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="047857")
    cell.alignment = Alignment(horizontal="center")
for column in range(1, sheet.max_column + 1):
    values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, min(sheet.max_row, 200) + 1)]
    sheet.column_dimensions[get_column_letter(column)].width = min(45, max(12, max(map(len, values)) + 2))
sheet.sheet_view.showGridLines = False
destination.parent.mkdir(parents=True, exist_ok=True)
workbook.save(destination)
print(f"Creado {destination} con {sheet.max_row - 1} productos")
